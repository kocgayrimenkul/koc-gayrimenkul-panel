# -*- coding: utf-8 -*-
import json
import io
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt

from .models import GelirKayit, KaporaKayit, Gider, GIDER_KATEGORI, ODEME_YONTEMI, OFİS_CHOICES


def _is_admin(user):
    try:
        ep = user.employee_profile.first()
        return user.is_superuser or (ep and ep.role in ('admin', 'manager'))
    except Exception:
        return user.is_superuser


def _personeller():
    from apps.authentication.models import CustomUser
    return CustomUser.objects.filter(is_active=True).order_by('first_name')


AY_ADLARI = ['', 'Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
             'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']


# ─── GELİRLER SAYFASI ────────────────────────────────────────

@login_required(login_url='/login/')
def gelir_list(request):
    if not _is_admin(request.user):
        messages.error(request, "Bu sayfaya erişim yetkiniz yok.")
        return redirect('home')

    now   = timezone.now()
    ay    = int(request.GET.get('ay',  now.month))
    yil   = int(request.GET.get('yil', now.year))

    gelirler_beykent   = list(GelirKayit.objects.filter(ofis='beykent',   ay=ay, yil=yil).select_related('bulan', 'satan'))
    gelirler_fistiklik = list(GelirKayit.objects.filter(ofis='fistiklik', ay=ay, yil=yil).select_related('bulan', 'satan'))

    kaporalar_beykent   = list(KaporaKayit.objects.filter(ofis='beykent',   ay=ay, yil=yil).select_related('satan'))
    kaporalar_fistiklik = list(KaporaKayit.objects.filter(ofis='fistiklik', ay=ay, yil=yil).select_related('satan'))

    toplam_beykent   = sum((g.toplam or 0) for g in gelirler_beykent)
    toplam_fistiklik = sum((g.toplam or 0) for g in gelirler_fistiklik)
    toplam_kap_bey   = sum((k.kapora or 0) for k in kaporalar_beykent)
    toplam_kap_fis   = sum((k.kapora or 0) for k in kaporalar_fistiklik)

    n = max(len(gelirler_beykent), len(gelirler_fistiklik),
            len(kaporalar_beykent), len(kaporalar_fistiklik), 1)

    rows = []
    for i in range(n):
        rows.append({
            'sira': i + 1,
            'gb': gelirler_beykent[i]   if i < len(gelirler_beykent)   else None,
            'gf': gelirler_fistiklik[i] if i < len(gelirler_fistiklik) else None,
            'kf': kaporalar_fistiklik[i] if i < len(kaporalar_fistiklik) else None,
            'kb': kaporalar_beykent[i]   if i < len(kaporalar_beykent)   else None,
        })

    return render(request, 'muhasebe/gelir_list.html', {
        'ay': ay, 'yil': yil,
        'ay_adi': AY_ADLARI[ay],
        'yillar': range(now.year + 1, 2023, -1),
        'aylar': range(1, 13),
        'ay_adlari': AY_ADLARI,
        'rows': rows,
        'toplam_beykent':   toplam_beykent,
        'toplam_fistiklik': toplam_fistiklik,
        'toplam_kap_bey':   toplam_kap_bey,
        'toplam_kap_fis':   toplam_kap_fis,
        'toplam_genel':     toplam_beykent + toplam_fistiklik,
        'toplam_kapora':    toplam_kap_bey + toplam_kap_fis,
        'personeller': _personeller(),
    })


# ─── KAPORA EKLE (AJAX) ──────────────────────────────────────

@login_required(login_url='/login/')
@require_POST
def kapora_ekle(request):
    if not _is_admin(request.user):
        return JsonResponse({'error': 'Yetki yok'}, status=403)
    try:
        body  = json.loads(request.body)
        ofis  = body.get('ofis', '')
        satan_id = body.get('satan_id')
        yer   = body.get('yer', '').strip()
        kapora = Decimal(str(body.get('kapora', 0) or 0))
        ay    = int(body.get('ay', timezone.now().month))
        yil   = int(body.get('yil', timezone.now().year))

        if not ofis or not yer:
            return JsonResponse({'error': 'Ofis ve yer zorunlu'}, status=400)

        from apps.authentication.models import CustomUser
        satan = CustomUser.objects.filter(pk=satan_id).first() if satan_id else None

        k = KaporaKayit.objects.create(
            ofis=ofis, satan=satan, yer=yer, kapora=kapora,
            ay=ay, yil=yil, olusturan=request.user,
        )
        return JsonResponse({
            'success': True,
            'id': k.id,
            'ofis': k.get_ofis_display(),
            'satan': k.satan.get_full_name() if k.satan else '—',
            'yer': k.yer,
            'kapora': float(k.kapora),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ─── KAPORA → GELİR (AJAX) ───────────────────────────────────

@login_required(login_url='/login/')
@require_POST
def kapora_gerceklestir(request, pk):
    if not _is_admin(request.user):
        return JsonResponse({'error': 'Yetki yok'}, status=403)
    try:
        k    = get_object_or_404(KaporaKayit, pk=pk)
        body = json.loads(request.body)
        from apps.authentication.models import CustomUser

        bulan_id = body.get('bulan_id')
        satan_id = body.get('satan_id')
        gelir_tutari = Decimal(str(body.get('gelir', 0) or 0))

        bulan = CustomUser.objects.filter(pk=bulan_id).first() if bulan_id else None
        satan = CustomUser.objects.filter(pk=satan_id).first() if satan_id else k.satan

        g = GelirKayit.objects.create(
            ofis=k.ofis,
            bulan=bulan,
            satan=satan,
            yer=k.yer,
            gelir=gelir_tutari,
            kapora=k.kapora,
            ay=k.ay,
            yil=k.yil,
            olusturan=request.user,
        )
        k.delete()

        return JsonResponse({
            'success': True,
            'gelir_id': g.id,
            'toplam': float(g.toplam),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ─── GELİR / KAPORA SİL ──────────────────────────────────────

@login_required(login_url='/login/')
@require_POST
def gelir_sil(request, pk):
    if not _is_admin(request.user):
        return JsonResponse({'error': 'Yetki yok'}, status=403)
    get_object_or_404(GelirKayit, pk=pk).delete()
    return JsonResponse({'success': True})


@login_required(login_url='/login/')
@require_POST
def kapora_sil(request, pk):
    if not _is_admin(request.user):
        return JsonResponse({'error': 'Yetki yok'}, status=403)
    get_object_or_404(KaporaKayit, pk=pk).delete()
    return JsonResponse({'success': True})


# ─── GİDERLER ────────────────────────────────────────────────

@login_required(login_url='/login/')
def gider_list(request):
    if not _is_admin(request.user):
        messages.error(request, "Bu sayfaya erişim yetkiniz yok.")
        return redirect('home')

    qs = Gider.objects.select_related('personel', 'olusturan')
    kategori = request.GET.get('kategori', '')
    ay  = request.GET.get('ay', '')
    yil = request.GET.get('yil', str(timezone.now().year))
    q   = request.GET.get('q', '')

    if kategori: qs = qs.filter(kategori=kategori)
    if yil: qs = qs.filter(tarih__year=yil)
    if ay:  qs = qs.filter(tarih__month=ay)
    if q:   qs = qs.filter(aciklama__icontains=q)

    toplam = qs.aggregate(t=Sum('tutar'))['t'] or 0

    return render(request, 'muhasebe/gider_list.html', {
        'giderler': qs, 'toplam': toplam,
        'kategori': kategori, 'ay': ay, 'yil': yil, 'q': q,
        'yillar': range(timezone.now().year, 2023, -1),
        'kategoriler': GIDER_KATEGORI,
        'odeme_yontemleri': ODEME_YONTEMI,
    })


@login_required(login_url='/login/')
def gider_ekle(request):
    if not _is_admin(request.user):
        return redirect('home')
    if request.method == 'POST':
        try:
            from apps.authentication.models import CustomUser
            g = Gider(
                tarih=request.POST['tarih'], tutar=request.POST['tutar'],
                kategori=request.POST['kategori'],
                aciklama=request.POST.get('aciklama', ''),
                odeme_yontemi=request.POST.get('odeme_yontemi', 'nakit'),
                olusturan=request.user,
            )
            pid = request.POST.get('personel')
            if pid: g.personel = CustomUser.objects.filter(pk=pid).first()
            if 'belge' in request.FILES: g.belge = request.FILES['belge']
            g.save()
            messages.success(request, 'Gider eklendi.')
            return redirect('gider_list')
        except Exception as e:
            messages.error(request, f'Hata: {e}')
    from apps.authentication.models import CustomUser
    return render(request, 'muhasebe/gider_form.html', {
        'personeller': CustomUser.objects.filter(is_active=True).order_by('first_name'),
        'kategoriler': GIDER_KATEGORI, 'odeme_yontemleri': ODEME_YONTEMI,
        'form_title': 'Gider Ekle',
    })


@login_required(login_url='/login/')
def gider_duzenle(request, pk):
    if not _is_admin(request.user):
        return redirect('home')
    gider = get_object_or_404(Gider, pk=pk)
    if request.method == 'POST':
        try:
            from apps.authentication.models import CustomUser
            gider.tarih = request.POST['tarih']
            gider.tutar = request.POST['tutar']
            gider.kategori = request.POST['kategori']
            gider.aciklama = request.POST.get('aciklama', '')
            gider.odeme_yontemi = request.POST.get('odeme_yontemi', 'nakit')
            pid = request.POST.get('personel')
            gider.personel = CustomUser.objects.filter(pk=pid).first() if pid else None
            if 'belge' in request.FILES: gider.belge = request.FILES['belge']
            gider.save()
            messages.success(request, 'Gider güncellendi.')
            return redirect('gider_list')
        except Exception as e:
            messages.error(request, f'Hata: {e}')
    from apps.authentication.models import CustomUser
    return render(request, 'muhasebe/gider_form.html', {
        'gider': gider,
        'personeller': CustomUser.objects.filter(is_active=True).order_by('first_name'),
        'kategoriler': GIDER_KATEGORI, 'odeme_yontemleri': ODEME_YONTEMI,
        'form_title': 'Gider Düzenle',
    })


@login_required(login_url='/login/')
@require_POST
def gider_sil(request, pk):
    if not _is_admin(request.user):
        return JsonResponse({'error': 'Yetki yok'}, status=403)
    get_object_or_404(Gider, pk=pk).delete()
    return JsonResponse({'success': True})


# ─── DASHBOARD ───────────────────────────────────────────────

@login_required(login_url='/login/')
def dashboard(request):
    """Muhasebe genel bakış: KPI'lar, aylık gelir/gider grafiği, kategori kırılımı."""
    if not _is_admin(request.user):
        messages.error(request, "Bu sayfaya erişim yetkiniz yok.")
        return redirect('home')

    now = timezone.now()
    yil = int(request.GET.get('yil', now.year))
    ay  = int(request.GET.get('ay',  now.month))

    gelir_ay = GelirKayit.objects.filter(yil=yil, ay=ay).aggregate(t=Sum('toplam'))['t'] or Decimal('0')
    gider_ay = Gider.objects.filter(tarih__year=yil, tarih__month=ay).aggregate(t=Sum('tutar'))['t'] or Decimal('0')
    kapora_ay = KaporaKayit.objects.filter(yil=yil, ay=ay).aggregate(t=Sum('kapora'))['t'] or Decimal('0')
    net_ay = gelir_ay - gider_ay

    gelir_yil = GelirKayit.objects.filter(yil=yil).aggregate(t=Sum('toplam'))['t'] or Decimal('0')
    gider_yil = Gider.objects.filter(tarih__year=yil).aggregate(t=Sum('tutar'))['t'] or Decimal('0')
    net_yil = gelir_yil - gider_yil

    aylik_gelir = [0] * 12
    aylik_gider = [0] * 12
    aylik_kapora = [0] * 12

    for row in GelirKayit.objects.filter(yil=yil).values('ay').annotate(t=Sum('toplam')):
        aylik_gelir[row['ay'] - 1] = float(row['t'] or 0)
    for m in range(1, 13):
        t = Gider.objects.filter(tarih__year=yil, tarih__month=m).aggregate(t=Sum('tutar'))['t'] or 0
        aylik_gider[m - 1] = float(t)
    for row in KaporaKayit.objects.filter(yil=yil).values('ay').annotate(t=Sum('kapora')):
        aylik_kapora[row['ay'] - 1] = float(row['t'] or 0)

    kategori_qs = (Gider.objects
                   .filter(tarih__year=yil, tarih__month=ay)
                   .values('kategori')
                   .annotate(t=Sum('tutar'), n=Count('id'))
                   .order_by('-t'))
    kategori_map = dict(GIDER_KATEGORI)
    kategori_labels = [kategori_map.get(r['kategori'], r['kategori']) for r in kategori_qs]
    kategori_data   = [float(r['t'] or 0) for r in kategori_qs]

    ofis_qs = (GelirKayit.objects.filter(yil=yil, ay=ay)
               .values('ofis').annotate(t=Sum('toplam')))
    ofis_map = dict(OFİS_CHOICES)
    ofis_labels = [ofis_map.get(r['ofis'], r['ofis']) for r in ofis_qs]
    ofis_data   = [float(r['t'] or 0) for r in ofis_qs]

    son_gelirler = GelirKayit.objects.select_related('bulan', 'satan').order_by('-created_at')[:5]
    son_giderler = Gider.objects.select_related('personel').order_by('-tarih', '-created_at')[:5]

    return render(request, 'muhasebe/dashboard.html', {
        'ay': ay, 'yil': yil,
        'ay_adi': AY_ADLARI[ay],
        'yillar': range(now.year + 1, 2023, -1),
        'aylar': range(1, 13),
        'ay_adlari': AY_ADLARI,
        'gelir_ay': gelir_ay, 'gider_ay': gider_ay,
        'kapora_ay': kapora_ay, 'net_ay': net_ay,
        'gelir_yil': gelir_yil, 'gider_yil': gider_yil, 'net_yil': net_yil,
        'chart_labels': json.dumps([AY_ADLARI[i] for i in range(1, 13)]),
        'chart_gelir':  json.dumps(aylik_gelir),
        'chart_gider':  json.dumps(aylik_gider),
        'chart_kapora': json.dumps(aylik_kapora),
        'kategori_labels': json.dumps(kategori_labels),
        'kategori_data':   json.dumps(kategori_data),
        'ofis_labels': json.dumps(ofis_labels),
        'ofis_data':   json.dumps(ofis_data),
        'son_gelirler': son_gelirler,
        'son_giderler': son_giderler,
    })


# ─── EXCEL EXPORT ────────────────────────────────────────────

def _xlsx_response(filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _style_header(ws, ncols, row=1):
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center


@login_required(login_url='/login/')
def gelir_export(request):
    if not _is_admin(request.user):
        return HttpResponse('Yetki yok', status=403)
    from openpyxl import Workbook
    from openpyxl.styles import Font

    now = timezone.now()
    ay  = int(request.GET.get('ay',  now.month))
    yil = int(request.GET.get('yil', now.year))

    qs = GelirKayit.objects.filter(yil=yil, ay=ay).select_related('bulan', 'satan').order_by('ofis', 'created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = f'Gelirler {AY_ADLARI[ay]} {yil}'
    headers = ['Ofis', 'Bulan', 'Satan', 'Yer', 'Gelir (TL)', 'Kapora (TL)', 'Toplam (TL)', 'Ay', 'Yil', 'Kayit Tarihi']
    ws.append(headers)
    _style_header(ws, len(headers))

    for g in qs:
        ws.append([
            g.get_ofis_display(),
            g.bulan.get_full_name() if g.bulan else '',
            g.satan.get_full_name() if g.satan else '',
            g.yer,
            float(g.gelir or 0),
            float(g.kapora or 0),
            float(g.toplam or 0),
            g.ay, g.yil,
            g.created_at.strftime('%d.%m.%Y %H:%M') if g.created_at else '',
        ])

    total = qs.aggregate(t=Sum('toplam'))['t'] or 0
    ws.append([])
    last = ws.max_row + 1
    bold = Font(bold=True)
    ws.cell(row=last, column=6, value='TOPLAM:').font = bold
    ws.cell(row=last, column=7, value=float(total)).font = bold

    widths = [12, 22, 22, 30, 14, 14, 14, 8, 8, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    response = _xlsx_response(f'gelirler-{yil}-{ay:02d}.xlsx')
    wb.save(response)
    return response


@login_required(login_url='/login/')
def gider_export(request):
    if not _is_admin(request.user):
        return HttpResponse('Yetki yok', status=403)
    from openpyxl import Workbook
    from openpyxl.styles import Font

    qs = Gider.objects.select_related('personel', 'olusturan')
    kategori = request.GET.get('kategori', '')
    ay  = request.GET.get('ay', '')
    yil = request.GET.get('yil', str(timezone.now().year))
    q   = request.GET.get('q', '')

    if kategori: qs = qs.filter(kategori=kategori)
    if yil: qs = qs.filter(tarih__year=yil)
    if ay:  qs = qs.filter(tarih__month=ay)
    if q:   qs = qs.filter(aciklama__icontains=q)

    qs = qs.order_by('-tarih')

    wb = Workbook()
    ws = wb.active
    ws.title = f'Giderler {yil}'
    headers = ['Tarih', 'Kategori', 'Tutar (TL)', 'Odeme Yontemi', 'Personel', 'Aciklama', 'Belge', 'Kaydeden']
    ws.append(headers)
    _style_header(ws, len(headers))

    kategori_map = dict(GIDER_KATEGORI)
    odeme_map = dict(ODEME_YONTEMI)
    for g in qs:
        ws.append([
            g.tarih.strftime('%d.%m.%Y') if g.tarih else '',
            kategori_map.get(g.kategori, g.kategori),
            float(g.tutar or 0),
            odeme_map.get(g.odeme_yontemi, g.odeme_yontemi),
            g.personel.get_full_name() if g.personel else '',
            g.aciklama or '',
            g.belge.url if g.belge else '',
            g.olusturan.get_full_name() if g.olusturan else '',
        ])

    total = qs.aggregate(t=Sum('tutar'))['t'] or 0
    ws.append([])
    last = ws.max_row + 1
    bold = Font(bold=True)
    ws.cell(row=last, column=2, value='TOPLAM:').font = bold
    ws.cell(row=last, column=3, value=float(total)).font = bold

    widths = [12, 18, 14, 16, 22, 40, 30, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    fname = f'giderler-{yil}' + (f'-{ay.zfill(2)}' if ay else '') + '.xlsx'
    response = _xlsx_response(fname)
    wb.save(response)
    return response


@login_required(login_url='/login/')
def dashboard_export(request):
    """Yillik ozet: aylik gelir/gider/net."""
    if not _is_admin(request.user):
        return HttpResponse('Yetki yok', status=403)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    now = timezone.now()
    yil = int(request.GET.get('yil', now.year))

    wb = Workbook()
    ws = wb.active
    ws.title = f'Ozet {yil}'

    ws.append([f'{yil} YILI MUHASEBE OZETI'])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])
    headers = ['Ay', 'Gelir (TL)', 'Gider (TL)', 'Net (TL)', 'Kapora Bekleyen (TL)']
    ws.append(headers)
    _style_header(ws, len(headers), row=3)

    toplam_gelir = Decimal('0')
    toplam_gider = Decimal('0')
    toplam_kapora = Decimal('0')
    for m in range(1, 13):
        gelir = GelirKayit.objects.filter(yil=yil, ay=m).aggregate(t=Sum('toplam'))['t'] or Decimal('0')
        gider = Gider.objects.filter(tarih__year=yil, tarih__month=m).aggregate(t=Sum('tutar'))['t'] or Decimal('0')
        kapora = KaporaKayit.objects.filter(yil=yil, ay=m).aggregate(t=Sum('kapora'))['t'] or Decimal('0')
        ws.append([AY_ADLARI[m], float(gelir), float(gider), float(gelir - gider), float(kapora)])
        toplam_gelir += gelir
        toplam_gider += gider
        toplam_kapora += kapora

    last = ws.max_row + 1
    bold = Font(bold=True)
    ws.cell(row=last, column=1, value='TOPLAM').font = bold
    ws.cell(row=last, column=2, value=float(toplam_gelir)).font = bold
    ws.cell(row=last, column=3, value=float(toplam_gider)).font = bold
    ws.cell(row=last, column=4, value=float(toplam_gelir - toplam_gider)).font = bold
    ws.cell(row=last, column=5, value=float(toplam_kapora)).font = bold

    widths = [12, 18, 18, 18, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    response = _xlsx_response(f'muhasebe-ozet-{yil}.xlsx')
    wb.save(response)
    return response



# ─── PERSONEL KAZANÇ (READ-ONLY) ─────────────────────────────

# Komisyon oranları
PRIM_SATAN = Decimal('0.15')   # %15
PRIM_BULAN = Decimal('0.15')   # %15 (ikisi aynı kişiyse toplam %30 olur)


@login_required(login_url='/login/')
def personel_kazanc(request):
    """Giriş yapan personelin seçili aydaki kazancı.
    Admin/manager ?personel=ID parametresiyle başka personelin kazancını görebilir.
    Yalnız okuma — düzenle/sil yok."""
    user = request.user
    now  = timezone.now()
    ay   = int(request.GET.get('ay',  now.month))
    yil  = int(request.GET.get('yil', now.year))

    # Admin/manager isterse başka personeli görebilir
    from apps.authentication.models import CustomUser
    hedef = user
    personeller = []
    is_admin_user = _is_admin(user)
    if is_admin_user:
        personeller = list(CustomUser.objects.filter(is_active=True).order_by('first_name', 'username'))
        personel_id = request.GET.get('personel')
        if personel_id:
            secilen = CustomUser.objects.filter(pk=personel_id).first()
            if secilen:
                hedef = secilen

    from django.db.models import Q
    qs = (GelirKayit.objects
          .filter(ay=ay, yil=yil)
          .filter(Q(satan=hedef) | Q(bulan=hedef))
          .select_related('bulan', 'satan')
          .order_by('ofis', 'yer'))

    kayitlar = []
    toplam_kazanc = Decimal('0')
    for g in qs:
        toplam = g.toplam or Decimal('0')
        rol = []
        oran = Decimal('0')
        if g.bulan_id == hedef.id:
            rol.append('Bulan')
            oran += PRIM_BULAN
        if g.satan_id == hedef.id:
            rol.append('Satan')
            oran += PRIM_SATAN
        kazanc = (toplam * oran).quantize(Decimal('0.01'))
        toplam_kazanc += kazanc
        kayitlar.append({
            'id': g.id,
            'ofis': g.get_ofis_display(),
            'yer': g.yer,
            'toplam': toplam,
            'rol': ' + '.join(rol) if rol else '—',
            'oran_pct': int(oran * 100),
            'kazanc': kazanc,
        })

    return render(request, 'muhasebe/personel_kazanc.html', {
        'ay': ay, 'yil': yil,
        'ay_adi': AY_ADLARI[ay],
        'yillar': range(now.year + 1, 2023, -1),
        'aylar': range(1, 13),
        'kayitlar': kayitlar,
        'toplam_kazanc': toplam_kazanc,
        'kullanici_ad': hedef.get_full_name() or hedef.username,
        'prim_satan_pct': int(PRIM_SATAN * 100),
        'prim_bulan_pct': int(PRIM_BULAN * 100),
        'is_admin': is_admin_user,
        'personeller': personeller,
        'hedef_id': hedef.id,
    })
            'rol': ' + '.join(rol) if rol else '—',
            'oran_pct': int(oran * 100),
            'kazanc': kazanc,
        })

    return render(request, 'muhasebe/personel_kazanc.html', {
        'ay': ay, 'yil': yil,
        'ay_adi': AY_ADLARI[ay],
        'yillar': range(now.year + 1, 2023, -1),
        'aylar': range(1, 13),
        'kayitlar': kayitlar,
        'toplam_kazanc': toplam_kazanc,
        'kullanici_ad': hedef.get_full_name() or hedef.username,
        'prim_satan_pct': int(PRIM_SATAN * 100),
        'prim_bulan_pct': int(PRIM_BULAN * 100),
        'is_admin': is_admin_user,
        'personeller': personeller,
        'hedef_id': hedef.id,
    })
