# -*- coding: utf-8 -*-
"""
Netgsm Santral Entegrasyonu Views
Santral webhook'ları ve arama yönetimi için view'lar
"""

import json
import logging
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg, Sum, Max
from .models import Lead, CallLog
from .netgsm_service import NetgsmService
from .forms import LeadFilterForm

logger = logging.getLogger(__name__)

@csrf_exempt
@require_http_methods(["GET", "POST"])
def netgsm_webhook(request):
    """
    Netgsm webhook endpoint
    GET: Webhook doğrulama
    POST: Arama event'lerini işle
    """
    if request.method == 'GET':
        # Webhook doğrulama
        verify_token = request.GET.get('verify_token')
        challenge = request.GET.get('challenge')
        
        netgsm_service = NetgsmService()
        if verify_token == netgsm_service.webhook_secret:
            return HttpResponse(challenge)
        else:
            return HttpResponse('Unauthorized', status=401)
    
    elif request.method == 'POST':
        try:
            # Webhook signature doğrulama
            signature = request.headers.get('X-Netgsm-Signature', '')
            body = request.body.decode('utf-8')
            
            netgsm_service = NetgsmService()
            
            if not netgsm_service.verify_webhook_signature(body, signature):
                logger.warning("Invalid webhook signature")
                return JsonResponse({'error': 'Invalid signature'}, status=401)
            
            # Webhook data'sını parse et
            webhook_data = json.loads(body)
            
            # Call webhook'unu işle
            call_log = netgsm_service.process_call_webhook(webhook_data)
            
            return JsonResponse({
                'success': True,
                'call_log_id': call_log.id if call_log else None
            })
            
        except json.JSONDecodeError:
            logger.error("Invalid JSON in webhook")
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        
        except Exception as e:
            logger.error(f"Webhook processing error: {str(e)}")
            return JsonResponse({'error': 'Processing failed'}, status=500)

@login_required
def call_logs(request):
    """
    Arama geçmişi listesi
    """
    # Filtreleme
    call_type = request.GET.get('call_type', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    phone_search = request.GET.get('phone_search', '')
    
    # Base queryset
    calls = CallLog.objects.select_related('lead').order_by('-start_time')
    
    # Filtreler
    if call_type:
        calls = calls.filter(call_type=call_type)
    
    if status:
        calls = calls.filter(status=status)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            calls = calls.filter(start_time__date__gte=date_from_obj.date())
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            calls = calls.filter(start_time__date__lte=date_to_obj.date())
        except ValueError:
            pass
    
    if phone_search:
        calls = calls.filter(
            Q(phone_number__icontains=phone_search) |
            Q(lead__customer_phone__icontains=phone_search) |
            Q(lead__customer_name__icontains=phone_search)
        )
    
    # Sayfalama
    paginator = Paginator(calls, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # İstatistikler
    stats = {
        'total_calls': calls.count(),
        'answered_calls': calls.filter(status='answered').count(),
        'missed_calls': calls.filter(status='missed').count(),
        'busy_calls': calls.filter(status='busy').count(),
        'total_duration': calls.aggregate(Sum('duration'))['duration__sum'] or 0,
        'avg_duration': calls.aggregate(Avg('duration'))['duration__avg'] or 0,
    }
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'call_type': call_type,
        'status': status,
        'date_from': date_from,
        'date_to': date_to,
        'phone_search': phone_search,
    }
    
    return render(request, 'sales_process/call_logs.html', context)

@login_required
def call_statistics(request):
    """
    Arama istatistikleri sayfası
    """
    # Tarih aralığı
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    end_date = timezone.now()
    
    # Netgsm servisinden istatistikleri al
    netgsm_service = NetgsmService()
    stats = netgsm_service.get_call_statistics(start_date, end_date)
    
    # Günlük trend verileri
    daily_stats = CallLog.objects.filter(
        start_time__range=[start_date, end_date]
    ).extra(
        select={'day': 'DATE(start_time)'}
    ).values('day').annotate(
        total_calls=Count('id'),
        answered_calls=Count('id', filter=Q(status='answered')),
        missed_calls=Count('id', filter=Q(status='missed')),
        total_duration=Sum('duration')
    ).order_by('day')
    
    # Agent performansı (eğer agent_extension bilgisi varsa)
    agent_stats = CallLog.objects.filter(
        start_time__range=[start_date, end_date],
        agent_extension__isnull=False
    ).values('agent_extension').annotate(
        total_calls=Count('id'),
        answered_calls=Count('id', filter=Q(status='answered')),
        total_duration=Sum('duration'),
        avg_duration=Avg('duration')
    ).order_by('-total_calls')
    
    # En çok aranan numaralar
    top_numbers = CallLog.objects.filter(
        start_time__range=[start_date, end_date]
    ).values('phone_number').annotate(
        call_count=Count('id'),
        last_call=Max('start_time')
    ).order_by('-call_count')[:10]
    
    context = {
        'stats': stats,
        'daily_stats': list(daily_stats),
        'agent_stats': list(agent_stats),
        'top_numbers': list(top_numbers),
        'days': days,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'sales_process/call_statistics.html', context)

@login_required
@require_http_methods(["POST"])
def make_call(request, lead_id):
    """
    Lead'e arama başlat
    """
    print(f"[DEBUG] make_call called with method: {request.method}, lead_id: {lead_id}")
    print(f"[DEBUG] POST data: {request.POST}")
    print(f"[DEBUG] Request headers: {dict(request.headers)}")
    
    # NetGSM entegrasyonu kapalıysa mock response döndür
    from django.conf import settings
    import time
    if not getattr(settings, 'NETGSM_ENABLED', False):
        print("[DEBUG] NetGSM integration is disabled, returning mock success response")
        
        # Mock response için değişkenleri başlat
        stage_updated = False
        new_stage_name = None
        
        try:
            # Lead'i bul ve mock aşama geçişi yap
            lead = get_object_or_404(Lead, lead_id=lead_id)
            print(f"[DEBUG] MOCK: Processing lead {lead.customer_name} for stage transition")
            
            # Mock arama başarılı notu ekle
            from .models import LeadNote, SalesStage, StageTransition
            LeadNote.objects.create(
                lead=lead,
                content=f"MOCK Arama başlatıldı - Test modu aktif",
                note_type='call',
                created_by=request.user
            )
            
            # Eğer lead 'bilgi-verildi' aşamasındaysa 'ihtiyac_analizi'ne geç
            if (lead.current_stage and 
                lead.current_stage.slug == 'bilgi-verildi'):
                
                try:
                    ihtiyac_analizi_stage = SalesStage.objects.get(slug='ihtiyac-analizi')
                    old_stage = lead.current_stage
                    
                    print(f"[DEBUG] MOCK: Moving lead from {old_stage.name} to {ihtiyac_analizi_stage.name}")
                    
                    # Aşama geçişini kaydet
                    StageTransition.objects.create(
                        lead=lead,
                        from_stage=old_stage,
                        to_stage=ihtiyac_analizi_stage,
                        transition_type='automatic',
                        reason='MOCK başarılı arama sonrası otomatik geçiş (Test modu)',
                        performed_by=request.user
                    )
                    
                    # Lead'i güncelle
                    lead.current_stage = ihtiyac_analizi_stage
                    lead.stage_updated_at = timezone.now()
                    lead.save()
                    
                    # Sistem notu ekle
                    LeadNote.objects.create(
                        lead=lead,
                        content=f"MOCK Otomatik aşama geçişi: {old_stage.name} → {ihtiyac_analizi_stage.name} (Test modu)",
                        note_type='system',
                        created_by=request.user
                    )
                    
                    print(f"[DEBUG] MOCK: Lead {lead.id} successfully moved to {ihtiyac_analizi_stage.name}")
                    stage_updated = True
                    new_stage_name = ihtiyac_analizi_stage.name
                    
                except SalesStage.DoesNotExist:
                    print("[DEBUG] MOCK: İhtiyaç Analizi aşaması bulunamadı")
                    
            else:
                print(f"[DEBUG] MOCK: Lead is not in 'bilgi-verildi' stage, current stage: {lead.current_stage.slug if lead.current_stage else 'None'}")
                
        except Exception as e:
            print(f"[DEBUG] MOCK: Error during mock stage transition: {str(e)}")
            import traceback
            print(f"[DEBUG] MOCK: Full traceback: {traceback.format_exc()}")
        
        # Mock response döndür
        response_data = {
            'success': True,
            'message': 'MOCK Arama başlatıldı (NetGSM kapalı)',
            'call_id': f'mock_call_{int(time.time())}'
        }
        
        if stage_updated:
            response_data['stage_updated'] = True
            response_data['new_stage'] = new_stage_name
            response_data['message'] = f'MOCK Arama başarılı - Aşama güncellendi: {new_stage_name}'
            
        return JsonResponse(response_data)
    
    try:
        print(f"[DEBUG] Looking for lead with ID: {lead_id}")
        lead = get_object_or_404(Lead, lead_id=lead_id)
        print(f"[DEBUG] Lead found: {lead.customer_name}, phone: {lead.customer_phone}")
        
        agent_extension = request.POST.get('agent_extension')
        print(f"[DEBUG] Agent extension: {agent_extension}")
        
        print("[DEBUG] Initializing NetGSM service")
        netgsm_service = NetgsmService()
        print(f"[DEBUG] NetGSM service initialized")
        
        print(f"[DEBUG] Making outbound call to {lead.customer_phone} with extension {agent_extension}")
        result = netgsm_service.make_outbound_call(
            phone_number=lead.customer_phone,
            agent_extension=agent_extension
        )
        
        print(f"[DEBUG] NetGSM result: {result}")
        
        if result['success']:
            print("[DEBUG] Call initiated successfully")
            # Arama başlatıldı notu ekle
            from .models import LeadNote
            LeadNote.objects.create(
                lead=lead,
                note=f"Arama başlatıldı - Agent: {agent_extension or 'Belirtilmedi'}",
                note_type='call',
                created_by=request.user
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Arama başlatıldı',
                'call_id': result.get('call_id')
            })
        else:
            print(f"[DEBUG] Call failed: {result.get('error', 'Arama başlatılamadı')}")
            return JsonResponse({
                'success': False,
                'error': result.get('error', 'Arama başlatılamadı')
            })
            
    except Exception as e:
        print(f"[DEBUG] Exception occurred: {str(e)}")
        import traceback
        print(f"[DEBUG] Full traceback: {traceback.format_exc()}")
        logger.error(f"Error making call: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Bir hata oluştu'
        })

@login_required
def call_detail(request, call_id):
    """
    Arama detayları
    """
    call_log = get_object_or_404(CallLog, id=call_id)
    
    context = {
        'call_log': call_log,
    }
    
    return render(request, 'sales_process/call_detail.html', context)

@login_required
def agent_dashboard(request):
    """
    Agent dashboard - günlük arama özeti
    """
    today = timezone.now().date()
    
    # Bugünkü aramalar
    today_calls = CallLog.objects.filter(
        start_time__date=today
    )
    
    # Agent extension'ı varsa filtrele
    agent_extension = request.GET.get('extension')
    if agent_extension:
        today_calls = today_calls.filter(agent_extension=agent_extension)
    
    # İstatistikler
    stats = {
        'total_calls': today_calls.count(),
        'incoming_calls': today_calls.filter(call_type='incoming').count(),
        'outgoing_calls': today_calls.filter(call_type='outgoing').count(),
        'answered_calls': today_calls.filter(status='answered').count(),
        'missed_calls': today_calls.filter(status='missed').count(),
        'total_duration': today_calls.aggregate(Sum('duration'))['duration__sum'] or 0,
    }
    
    # Son aramalar
    recent_calls = today_calls.select_related('lead').order_by('-start_time')[:10]
    
    # Bekleyen görevler (kaçırılmış aramalardan)
    from .models import Task
    pending_tasks = Task.objects.filter(
        task_type='call',
        status='pending',
        assigned_to=request.user
    ).select_related('lead')[:5]
    
    context = {
        'stats': stats,
        'recent_calls': recent_calls,
        'pending_tasks': pending_tasks,
        'agent_extension': agent_extension,
        'today': today,
    }
    
    return render(request, 'sales_process/agent_dashboard.html', context)

@login_required
@require_http_methods(["POST"])
def update_call_notes(request, call_id):
    """
    Arama notlarını güncelle
    """
    try:
        call_log = get_object_or_404(CallLog, id=call_id)
        notes = request.POST.get('notes', '')
        
        call_log.notes = notes
        call_log.save()
        
        # Lead'e de not ekle
        if call_log.lead and notes:
            from .models import LeadNote
            LeadNote.objects.create(
                lead=call_log.lead,
                title="Arama Notu",
                content=f"Arama notu: {notes}",
                note_type='call',
                created_by=request.user
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Notlar güncellendi'
        })
        
    except Exception as e:
        logger.error(f"Error updating call notes: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Notlar güncellenemedi'
        })

@login_required
def export_call_logs(request):
    """
    Arama geçmişini Excel olarak dışa aktar
    """
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.utils import get_column_letter
    
    # Filtreleri uygula
    calls = CallLog.objects.select_related('lead').order_by('-start_time')
    
    # Tarih filtresi
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            calls = calls.filter(start_time__date__gte=date_from_obj.date())
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            calls = calls.filter(start_time__date__lte=date_to_obj.date())
        except ValueError:
            pass
    
    # Excel dosyası oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arama Geçmişi"
    
    # Başlıklar
    headers = [
        'Tarih/Saat', 'Telefon', 'Müşteri Adı', 'Arama Tipi', 
        'Durum', 'Süre (sn)', 'Agent', 'Notlar'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Veriler
    for row, call in enumerate(calls[:1000], 2):  # Max 1000 kayıt
        ws.cell(row=row, column=1, value=call.start_time.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=2, value=call.phone_number)
        ws.cell(row=row, column=3, value=call.lead.customer_name if call.lead else 'Bilinmiyor')
        ws.cell(row=row, column=4, value=call.get_call_type_display())
        ws.cell(row=row, column=5, value=call.get_status_display())
        ws.cell(row=row, column=6, value=call.duration)
        ws.cell(row=row, column=7, value=call.agent_extension or '')
        ws.cell(row=row, column=8, value=call.notes or '')
    
    # Sütun genişliklerini ayarla
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Response oluştur
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="arama_gecmisi_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
@require_http_methods(["POST"])
def hangup_call(request):
    """
    Çağrı sonlandırma
    """
    try:
        unique_id = request.POST.get('unique_id')
        
        if not unique_id:
            return JsonResponse({
                'success': False,
                'error': 'Unique ID gerekli'
            })
        
        netgsm_service = NetgsmService()
        result = netgsm_service.hangup_call(unique_id)
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"Error hanging up call: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Bir hata oluştu'
        })

@login_required
@require_http_methods(["POST"])
def mute_call(request):
    """
    Çağrıyı sessize alma/açma
    """
    try:
        unique_id = request.POST.get('unique_id')
        direction = request.POST.get('direction', 'all')  # all/in/out
        state = request.POST.get('state', 'mute')  # mute/unmute
        
        if not unique_id:
            return JsonResponse({
                'success': False,
                'error': 'Unique ID gerekli'
            })
        
        netgsm_service = NetgsmService()
        result = netgsm_service.mute_call(unique_id, direction, state)
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"Error muting call: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Bir hata oluştu'
        })

@login_required
@require_http_methods(["POST"])
def link_calls(request):
    """
    İki numarayı birbirine bağlama
    """
    try:
        caller_number = request.POST.get('caller_number')
        called_number = request.POST.get('called_number')
        ring_timeout = int(request.POST.get('ring_timeout', 20))
        originate_order = request.POST.get('originate_order', 'if')
        
        if not caller_number or not called_number:
            return JsonResponse({
                'success': False,
                'error': 'Caller ve called numaraları gerekli'
            })
        
        netgsm_service = NetgsmService()
        result = netgsm_service.link_calls(
            caller_number=caller_number,
            called_number=called_number,
            ring_timeout=ring_timeout,
            originate_order=originate_order
        )
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"Error linking calls: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Bir hata oluştu'
        })

@login_required
def queue_stats(request):
    """
    Kuyruk istatistikleri
    """
    try:
        queue_name = request.GET.get('queue_name')
        
        if not queue_name:
            return JsonResponse({
                'success': False,
                'error': 'Kuyruk adı gerekli'
            })
        
        netgsm_service = NetgsmService()
        result = netgsm_service.get_queue_stats(queue_name)
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"Error getting queue stats: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Bir hata oluştu'
        })

@login_required
@require_http_methods(["POST"])
def agent_login(request):
    """
    Dahiliyi kuyruğa ekleme
    """
    try:
        queue_name = request.POST.get('queue_name')
        extension = request.POST.get('extension')
        paused = int(request.POST.get('paused', 1))
        penalty = int(request.POST.get('penalty', 1))
        
        if not queue_name or not extension:
            return JsonResponse({
                'success': False,
                'error': 'Kuyruk adı ve dahili numarası gerekli'
            })
        
        netgsm_service = NetgsmService()
        result = netgsm_service.agent_login(
            queue_name=queue_name,
            extension=extension,
            paused=paused,
            penalty=penalty
        )
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"Error agent login: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Bir hata oluştu'
        })

@login_required
@require_http_methods(["POST"])
def agent_logout(request):
    """
    Dahiliyi kuyruktan çıkarma
    """
    try:
        queue_name = request.POST.get('queue_name')
        extension = request.POST.get('extension')
        
        if not queue_name or not extension:
            return JsonResponse({
                'success': False,
                'error': 'Kuyruk adı ve dahili numarası gerekli'
            })
        
        netgsm_service = NetgsmService()
        result = netgsm_service.agent_logout(
            queue_name=queue_name,
            extension=extension
        )
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"Error agent logout: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Bir hata oluştu'
        })

@login_required
@require_http_methods(["POST"])
def agent_pause(request):
    """
    Dahiliyi molaya alma/çıkarma
    """
    try:
        queue_name = request.POST.get('queue_name')
        extension = request.POST.get('extension')
        paused = int(request.POST.get('paused', 1))
        reason = request.POST.get('reason', 'CRM')
        
        if not queue_name or not extension:
            return JsonResponse({
                'success': False,
                'error': 'Kuyruk adı ve dahili numarası gerekli'
            })
        
        netgsm_service = NetgsmService()
        result = netgsm_service.agent_pause(
            queue_name=queue_name,
            extension=extension,
            paused=paused,
            reason=reason
        )
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"Error agent pause: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Bir hata oluştu'
        })

@login_required
@require_http_methods(["POST"])
def add_external_number_to_queue(request):
    """
    Kuyruğa dış numara ekleme
    """
    try:
        queue_name = request.POST.get('queue_name')
        phone_number = request.POST.get('phone_number')
        penalty = int(request.POST.get('penalty', 1))
        
        if not queue_name or not phone_number:
            return JsonResponse({
                'success': False,
                'error': 'Kuyruk adı ve telefon numarası gerekli'
            })
        
        netgsm_service = NetgsmService()
        result = netgsm_service.add_external_number_to_queue(
            queue_name=queue_name,
            phone_number=phone_number,
            penalty=penalty
        )
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"Error adding external number to queue: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Bir hata oluştu'
        })

@login_required
@require_http_methods(["POST"])
def dynamic_redirect(request):
    """
    Dinamik yönlendirme
    """
    try:
        called_number = request.POST.get('called_number')
        redirect_menu = request.POST.get('redirect_menu')
        redirect_type = request.POST.get('redirect_type', 'ivr')
        ring_timeout = int(request.POST.get('ring_timeout', 20))
        
        if not called_number or not redirect_menu:
            return JsonResponse({
                'success': False,
                'error': 'Aranacak numara ve yönlendirme menüsü gerekli'
            })
        
        netgsm_service = NetgsmService()
        result = netgsm_service.dynamic_redirect(
            called_number=called_number,
            redirect_menu=redirect_menu,
            redirect_type=redirect_type,
            ring_timeout=ring_timeout
        )
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"Error dynamic redirect: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Bir hata oluştu'
        })

@login_required
def netgsm_management(request):
    """
    NetGSM Santral Yönetim Paneli
    """
    context = {
        'page_title': 'NetGSM Santral Yönetimi',
    }
    
    return render(request, 'sales_process/netgsm_management.html', context)